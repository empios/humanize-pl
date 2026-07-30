"""End-to-end flow over one column of an .xlsx sheet.

Intended for a sheet of AI-drafted answers: point the flow at the column that
holds them and it appends the diagnosis, the gate verdict and the regeneration
constraints as new columns next to each row.
"""

from __future__ import annotations

import json
import unicodedata
from pathlib import Path
from typing import Any

from .base import FlowSettings, ItemOutcome, run_all_layers, summarise

# Appended to the right of the existing data, in this order.
OUTPUT_COLUMNS = [
    "sygnał AI",
    "do przeglądu",
    "znaleziska",
    "rodziny",
    "ograniczenia do regeneracji",
]
REWRITE_COLUMN = "tekst po redakcji"


def _require_openpyxl():
    try:
        import openpyxl  # type: ignore
    except ImportError as exc:  # pragma: no cover - depends on the install
        raise RuntimeError(
            "Obsługa .xlsx wymaga openpyxl. Zainstaluj: pip install -e '.[xlsx]'"
        ) from exc
    return openpyxl


def resolve_column(sheet, spec: str, *, header_row: int | None) -> int:
    """Accept a column letter, a 1-based index, or a header cell's text.

    Header matching is case- and whitespace-insensitive so "Odpowiedź AI" and
    "odpowiedz ai " both resolve, which is what a hand-made sheet looks like.
    """
    from openpyxl.utils import column_index_from_string  # type: ignore

    spec = spec.strip()
    if spec.isdigit():
        return int(spec)
    if spec.isalpha() and len(spec) <= 3:
        try:
            return column_index_from_string(spec.upper())
        except ValueError:
            pass

    if header_row:
        wanted = _normalise(spec)
        for cell in sheet[header_row]:
            if cell.value is not None and _normalise(str(cell.value)) == wanted:
                return cell.column

    raise ValueError(
        f"Nie znaleziono kolumny „{spec}”. Podaj literę (np. D), numer (np. 4) "
        "albo nagłówek, wskazując --header-row."
    )


def run_xlsx_flow(
    input_path: Path,
    output_path: Path,
    *,
    column: str,
    settings: FlowSettings,
    sheet_name: str | None = None,
    header_row: int | None = 1,
    report_path: Path | None = None,
    on_item=None,
) -> dict[str, Any]:
    openpyxl = _require_openpyxl()

    workbook = openpyxl.load_workbook(str(input_path))
    sheet = workbook[sheet_name] if sheet_name else workbook.active
    column_index = resolve_column(sheet, column, header_row=header_row)

    first_column = sheet.max_column + 1
    headers = list(OUTPUT_COLUMNS)
    if settings.rewrite:
        headers.append(REWRITE_COLUMN)
    if header_row:
        for offset, header in enumerate(headers):
            sheet.cell(row=header_row, column=first_column + offset, value=header)

    session = settings.session() if settings.rewrite else None
    start_row = (header_row + 1) if header_row else 1
    outcomes: list[ItemOutcome] = []

    for row_index in range(start_row, sheet.max_row + 1):
        value = sheet.cell(row=row_index, column=column_index).value
        text = str(value).strip() if value is not None else ""
        if not text:
            continue

        name = f"wiersz {row_index}"
        try:
            outcome, _verdict = run_all_layers(
                text, name=name, settings=settings, session=session
            )
        except Exception as exc:
            outcome = ItemOutcome(name=name, status="failed", error=f"{type(exc).__name__}: {exc}")
            outcomes.append(outcome)
            if on_item is not None:
                on_item(outcome)
            continue

        values = [
            outcome.signal_after,
            "TAK" if outcome.needs_review else "nie",
            outcome.findings_before,
            "; ".join(outcome.families),
            "\n".join(f"• {item}" for item in outcome.constraints),
        ]
        if settings.rewrite:
            values.append(outcome.text_out or text)
        for offset, cell_value in enumerate(values):
            sheet.cell(row=row_index, column=first_column + offset, value=cell_value)

        outcomes.append(outcome)
        if on_item is not None:
            on_item(outcome)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(str(output_path))

    payload = {
        "flow": "xlsx",
        "input_path": str(input_path),
        "output_path": str(output_path),
        "sheet": sheet.title,
        "source_column": column,
        "source_column_index": column_index,
        "settings": {
            "mode": settings.mode.value,
            "engine": settings.engine.value,
            "rewrite": settings.rewrite,
            "require_anchor": settings.require_anchor,
        },
        "summary": summarise(outcomes),
        "rows": [item.to_json() for item in outcomes],
    }
    if report_path is not None:
        report_path.parent.mkdir(parents=True, exist_ok=True)
        report_path.write_text(
            json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
        )
    return payload


def _normalise(value: str) -> str:
    """Fold case, whitespace and Polish diacritics.

    Hand-made sheets are typed without diacritics as often as with them, so
    "Odpowiedź AI" and "odpowiedz ai" have to resolve to the same column.
    """
    folded = unicodedata.normalize("NFKD", value)
    stripped = "".join(char for char in folded if not unicodedata.combining(char))
    # ł/Ł has no decomposition, so NFKD leaves it alone.
    stripped = stripped.replace("ł", "l").replace("Ł", "L")
    return " ".join(stripped.split()).casefold()
