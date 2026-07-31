"""Tests for the end-to-end flows."""

from __future__ import annotations

import json

import pytest

from humanize_pl.config import Engine, Mode
from humanize_pl.flows import FlowSettings, run_all_layers, run_docx_flow, run_xlsx_flow

openpyxl = pytest.importorskip("openpyxl")

# The flow now defaults to the hybrid neural stack, which costs ~35 s to load.
# Tests pin the engine to basic; engine selection is covered separately.
BASIC = FlowSettings(mode=Mode.standard, engine=Engine.basic)
BASIC_NO_REWRITE = FlowSettings(engine=Engine.basic, rewrite=False)

AI_TEXT = (
    "Podporządkowanie pracownika jest jedną z najważniejszych cech stosunku pracy.\n"
    "Warto wskazać, że podporządkowanie nie oznacza całkowitej zależności. "
    "Szczególne znaczenie ma tutaj kierownictwo pracodawcy.\n"
    "Z jednej strony umożliwia to organizowanie pracy. Podsumowując, ma to duże znaczenie."
)


def write_docx(path, text):
    from docx import Document

    document = Document()
    for paragraph in text.split("\n"):
        if paragraph.strip():
            document.add_paragraph(paragraph)
    document.save(str(path))


def write_xlsx(path, rows, *, header=("ID", "Odpowiedź AI")):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(list(header))
    for row in rows:
        sheet.append(list(row))
    workbook.save(str(path))


def test_layer_status_reports_what_actually_loaded() -> None:
    from humanize_pl.flows.base import layer_status

    status = layer_status(BASIC.session())

    # Morfeusz backs detection and rewriting alike; Stanza never backs detection.
    assert status["detection"]["stanza"] == "not_used"
    assert status["detection"]["morfeusz"] in {"ready", "unavailable"}
    assert status["rewrite"]["engine_requested"] == "basic"
    assert "morfeusz" in status["rewrite"]


def test_layer_status_marks_a_skipped_rewrite() -> None:
    from humanize_pl.flows.base import layer_status

    assert layer_status(None)["rewrite"] == {"skipped": True}


def test_flow_report_records_the_layer_status(tmp_path) -> None:
    source = tmp_path / "in"
    source.mkdir()
    write_docx(source / "a.docx", AI_TEXT)

    payload = run_docx_flow(source, tmp_path / "out", settings=BASIC_NO_REWRITE)

    assert payload["layers"]["detection"]["stanza"] == "not_used"


def test_run_all_layers_reports_signal_before_and_after() -> None:
    outcome, verdict = run_all_layers(
        AI_TEXT, name="t", settings=BASIC
    )

    assert outcome.signal_before > 0
    assert outcome.findings_before > 0
    assert outcome.text_out is not None
    assert verdict.prompt_constraints


def test_no_rewrite_leaves_the_text_untouched() -> None:
    outcome, _ = run_all_layers(AI_TEXT, name="t", settings=BASIC_NO_REWRITE)

    assert outcome.text_out == AI_TEXT
    assert outcome.changes_applied == 0
    assert outcome.signal_after == outcome.signal_before


def test_docx_flow_writes_documents_reports_and_summary(tmp_path) -> None:
    source = tmp_path / "in"
    source.mkdir()
    write_docx(source / "opinia.docx", AI_TEXT)
    output = tmp_path / "out"

    payload = run_docx_flow(source, output, settings=BASIC)

    assert payload["summary"]["ok"] == 1
    assert (output / "opinia_humanized.docx").exists()
    assert (output / "summary.csv").exists()
    assert (output / "details" / "opinia.json").exists()

    detail = json.loads((output / "details" / "opinia.json").read_text(encoding="utf-8"))
    assert detail["findings"]
    assert detail["gate"]["prompt_constraints"]


def test_docx_flow_needs_at_least_one_document(tmp_path) -> None:
    empty = tmp_path / "empty"
    empty.mkdir()

    with pytest.raises(FileNotFoundError):
        run_docx_flow(empty, tmp_path / "out", settings=BASIC)


def test_docx_flow_keeps_going_after_a_broken_file(tmp_path) -> None:
    source = tmp_path / "in"
    source.mkdir()
    write_docx(source / "dobry.docx", AI_TEXT)
    (source / "zepsuty.docx").write_bytes(b"not a docx")

    payload = run_docx_flow(source, tmp_path / "out", settings=BASIC_NO_REWRITE)

    assert payload["summary"]["ok"] == 1
    assert payload["summary"]["failed"] == 1


@pytest.mark.parametrize("column", ["B", "2", "Odpowiedź AI", "odpowiedz ai"])
def test_xlsx_column_can_be_named_by_letter_index_or_header(tmp_path, column: str) -> None:
    source = tmp_path / "in.xlsx"
    write_xlsx(source, [("1", AI_TEXT)])
    output = tmp_path / "out.xlsx"

    payload = run_xlsx_flow(
        source, output, column=column, settings=BASIC_NO_REWRITE
    )

    assert payload["source_column_index"] == 2
    assert payload["summary"]["ok"] == 1


def test_xlsx_flow_appends_result_columns_without_touching_the_source(tmp_path) -> None:
    source = tmp_path / "in.xlsx"
    write_xlsx(source, [("1", AI_TEXT)])
    output = tmp_path / "out.xlsx"

    run_xlsx_flow(source, output, column="B", settings=BASIC)

    sheet = openpyxl.load_workbook(str(output)).active
    headers = [cell.value for cell in sheet[1]]
    assert headers[:2] == ["ID", "Odpowiedź AI"]
    assert "sygnał AI" in headers
    assert "tekst po redakcji" in headers
    assert sheet.cell(row=2, column=2).value == AI_TEXT

    verdict_column = headers.index("do przeglądu") + 1
    assert sheet.cell(row=2, column=verdict_column).value == "TAK"


def test_xlsx_flow_skips_blank_cells(tmp_path) -> None:
    source = tmp_path / "in.xlsx"
    write_xlsx(source, [("1", AI_TEXT), ("2", ""), ("3", None)])

    payload = run_xlsx_flow(
        source, tmp_path / "out.xlsx", column="B", settings=BASIC_NO_REWRITE
    )

    assert payload["summary"]["items"] == 1


def test_xlsx_flow_reports_an_empty_column_instead_of_writing_nothing(tmp_path) -> None:
    """Silently copying the file is the failure mode this project started with."""
    source = tmp_path / "in.xlsx"
    write_xlsx(source, [("1", AI_TEXT)], header=("ID", "Odpowiedź AI", "Pusta"))
    output = tmp_path / "out.xlsx"

    with pytest.raises(ValueError, match="nie ma żadnych niepustych komórek"):
        run_xlsx_flow(source, output, column="C", settings=BASIC_NO_REWRITE)

    assert not output.exists()


def test_xlsx_flow_names_the_other_sheets_when_the_active_one_is_empty(tmp_path) -> None:
    workbook = openpyxl.Workbook()
    workbook.active.title = "Podsumowanie"
    workbook.active["A1"] = "nic"
    data = workbook.create_sheet("Dane")
    data.append(["ID", "Odpowiedź AI"])
    data.append(["1", AI_TEXT])
    source = tmp_path / "in.xlsx"
    workbook.save(str(source))

    with pytest.raises(ValueError, match="Dane"):
        run_xlsx_flow(source, tmp_path / "out.xlsx", column="B", settings=BASIC_NO_REWRITE)

    payload = run_xlsx_flow(
        source,
        tmp_path / "out.xlsx",
        column="B",
        sheet_name="Dane",
        settings=BASIC_NO_REWRITE,
    )
    assert payload["summary"]["ok"] == 1


def test_xlsx_flow_reads_cached_formula_results(tmp_path) -> None:
    """A column of answers pulled from elsewhere holds formulas, not text."""
    from openpyxl.worksheet.formula import ArrayFormula  # noqa: F401  (import guard)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["ID", "Odpowiedź AI"])
    sheet.append(["1", "=A2"])
    source = tmp_path / "in.xlsx"
    workbook.save(str(source))

    # No Excel has opened this file, so there is no cached value; the raw
    # fallback keeps the row from vanishing.
    payload = run_xlsx_flow(
        source, tmp_path / "out.xlsx", column="B", settings=BASIC_NO_REWRITE
    )

    assert payload["summary"]["items"] == 1


def test_xlsx_output_columns_land_beside_the_data_not_further_out(tmp_path) -> None:
    source = tmp_path / "in.xlsx"
    write_xlsx(source, [("1", AI_TEXT)])
    output = tmp_path / "out.xlsx"

    run_xlsx_flow(source, output, column="B", settings=BASIC_NO_REWRITE)

    sheet = openpyxl.load_workbook(str(output)).active
    assert sheet.cell(row=1, column=3).value == "sygnał AI"


def test_xlsx_flow_rejects_an_unknown_column(tmp_path) -> None:
    source = tmp_path / "in.xlsx"
    write_xlsx(source, [("1", AI_TEXT)])

    with pytest.raises(ValueError, match="Nie znaleziono kolumny"):
        run_xlsx_flow(source, tmp_path / "out.xlsx", column="Brak", settings=BASIC)


def test_xlsx_flow_without_headers_starts_at_the_first_row(tmp_path) -> None:
    workbook = openpyxl.Workbook()
    workbook.active.append(["1", AI_TEXT])
    source = tmp_path / "in.xlsx"
    workbook.save(str(source))

    payload = run_xlsx_flow(
        source,
        tmp_path / "out.xlsx",
        column="B",
        header_row=None,
        settings=BASIC_NO_REWRITE,
    )

    assert payload["summary"]["items"] == 1
